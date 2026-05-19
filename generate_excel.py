import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()

# Aba 1: Projetos
ws1 = wb.active
ws1.title = "Projetos"
cols1 = ["Data", "Projeto", "Tipo", "Tempo Manual (min)", "Tempo Sistema (min)", "Economia (%)", "Status", "Erros Encontrados", "Observações"]
ws1.append(cols1)

# Aba 2: Erros
ws2 = wb.create_sheet("Erros")
cols2 = ["#", "Data", "Projeto", "Etapa", "Descrição", "Impacto", "Status", "Solução/Workaround"]
ws2.append(cols2)

# Aba 3: Métricas
ws3 = wb.create_sheet("Métricas")
cols3 = ["Métrica", "Valor Esperado/Meta", "Valor Atual"]
ws3.append(cols3)
metrics = [
    ["Projetos testados", "10", "0"],
    ["Projetos OK", "10", "0"],
    ["Taxa sucesso", ">= 80%", "0%"],
    ["Tempo médio antes", "-", "-"],
    ["Tempo médio depois", "-", "-"],
    ["Economia tempo", ">= 30%", "0%"],
    ["Erros críticos", "0", "0"],
    ["Erros altos", "-", "0"],
    ["Erros médios", "-", "0"],
    ["Erros baixos", "-", "0"],
]
for m in metrics:
    ws3.append(m)

# Aba 4: Feedback Produção
ws4 = wb.create_sheet("Feedback Produção")
cols4 = ["Data", "Feedback", "Categoria"]
ws4.append(cols4)

# Formatting headers
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
alignment = Alignment(horizontal="center", vertical="center")

for ws in wb.worksheets:
    for col in range(1, len(ws[1]) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        # Set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

wb.save("validacao_marcenai_dluxury.xlsx")
